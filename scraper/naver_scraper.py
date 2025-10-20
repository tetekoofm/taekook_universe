import os
import time
import pandas as pd
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def scrape_naver_blog(blog_url):
    # --- Selenium setup ---
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(options=options)

    driver.get(blog_url)
    time.sleep(3)

    # --- Scroll gradually to load all posts ---
    SCROLL_PAUSE = 2
    max_attempts = 20
    attempt = 0
    last_count = 0

    while attempt < max_attempts:
        driver.execute_script("window.scrollBy(0, window.innerHeight);")
        time.sleep(SCROLL_PAUSE)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        posts = soup.find_all("div", class_="card__reUkU")

        if len(posts) == last_count:
            attempt += 1
        else:
            last_count = len(posts)
            attempt = 0

    print(f"Total posts found: {len(posts)}")

    # --- Extract post data ---
    post_data = []
    for post in posts:
        try:
            date_span = post.find("span", class_="time__Uowk3")
            date = date_span.get_text(strip=True).replace(" ", "").replace(".", "-").strip("-") if date_span else "No Date"

            img_tag = post.find("img")
            image = img_tag.get("data-src") if img_tag and img_tag.get("data-src") else ""
            title = img_tag.get("alt", "No Title") if img_tag else "No Title"

            link_tag = post.find("a", class_="link__Vqgpc")
            link = link_tag["href"] if link_tag else ""

            description_tag = post.find("div", class_="text_area__mOuKZ")
            description = description_tag.get_text(strip=True) if description_tag else ""

            text_combined = f"{title} {description}".lower()
            if "taekook" in text_combined or (("v" in text_combined or "taehyung" in text_combined) and "jungkook" in text_combined):
                artist = "TaeKook"
            elif "jungkook" in text_combined or "정국" in text_combined:
                artist = "Jungkook"
            elif "v" in text_combined or "태형" in text_combined:
                artist = "Taehyung"
            else:
                artist = "Unknown"

            post_data.append([date, artist, title, image, description, link])
            print(f"Added post: {title} → {artist}")

        except Exception as e:
            print(f"Error processing post: {e}")

    driver.quit()

    # --- Excel setup ---
    main_folder = os.path.dirname(os.path.abspath(__file__))
    main_excel = os.path.join(os.path.dirname(main_folder), "taekook_universe.xlsx")
    sheet_name = "In The News"

    df_new = pd.DataFrame(post_data, columns=["date", "artist", "title", "image", "description", "link"])

    # --- Load existing workbook safely ---
    if os.path.exists(main_excel):
        wb = load_workbook(main_excel)
        if sheet_name in wb.sheetnames:
            try:
                old_df = pd.read_excel(main_excel, sheet_name=sheet_name)
                if "link" in old_df.columns:
                    old_links = set(old_df["link"].dropna().tolist())
                    df_filtered = df_new[~df_new["link"].isin(old_links)]
                else:
                    df_filtered = df_new
                    print("⚠️ Existing sheet missing 'link' column — adding all new posts.")

                if not df_filtered.empty:
                    updated_df = pd.concat([old_df, df_filtered], ignore_index=True)
                    with pd.ExcelWriter(main_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        updated_df.to_excel(writer, index=False, sheet_name=sheet_name)
                    print(f"✅ Added {len(df_filtered)} new posts.")
                else:
                    print("No new posts found.")
            except Exception as e:
                print(f"⚠️ Could not read 'In The News' properly: {e}")
                with pd.ExcelWriter(main_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, index=False, sheet_name=sheet_name)
                print(f"✅ Created or updated sheet safely: {sheet_name}")
        else:
            # Create new sheet without deleting others
            with pd.ExcelWriter(main_excel, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
                df_new.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"✅ Added new sheet '{sheet_name}' to existing workbook.")
    else:
        df_new.to_excel(main_excel, index=False, sheet_name=sheet_name)
        print(f"✅ Created new workbook: {main_excel}")

    print(f"Scraping completed! {len(df_new)} posts fetched.")

# --- Run ---
if __name__ == "__main__":
    blog_url = "https://m.blog.naver.com/uktaekook"
    scrape_naver_blog(blog_url)
